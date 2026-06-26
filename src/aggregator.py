# -*- coding: utf-8 -*-
"""销售单 Excel 汇总核心逻辑。"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Callable, List, Optional, Union

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from config import EXTERNAL_TEMPLATE, INTERNAL_TEMPLATE, TEMPLATES, TemplateConfig

# 进度回调：(进度 0~1, 状态说明)
ProgressCallback = Callable[[float, str], None]


@dataclass
class AggregateResult:
    """汇总结果。"""

    data: pd.DataFrame
    raw_row_count: int
    result_row_count: int
    template_id: str
    template_label: str
    dimension_label: str


class AggregateError(Exception):
    """汇总过程中的业务错误。"""


def _noop_progress(percent: float, message: str) -> None:
    """空进度回调。"""


# 常见列名变体 -> 标准列名（仅当标准列名尚不存在时重命名）
COLUMN_ALIASES = {
    "金额（元）": "金额",
    "金额(元)": "金额",
    "金额（含税）": "金额",
    "数量（件）": "数量",
    "数量(件)": "数量",
}


def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """将表头常见变体统一为标准列名，便于模板识别与汇总。"""
    rename_map = {}
    existing = set(df.columns)
    for column in df.columns:
        target = COLUMN_ALIASES.get(str(column).strip())
        if target and target not in existing and column not in rename_map:
            rename_map[column] = target
            existing.add(target)
    if not rename_map:
        return df
    return df.rename(columns=rename_map)


def _read_excel_with_progress(path: Path, callback: ProgressCallback) -> pd.DataFrame:
    """读取 Excel 并在读取过程中汇报进度。"""
    suffix = path.suffix.lower()

    if suffix == ".xls":
        callback(0.1, "正在读取 Excel 文件...")
        df = pd.read_excel(path, sheet_name=0)
        callback(0.55, f"读取完成，共 {len(df)} 行")
        return df

    callback(0.05, "正在打开 Excel 文件...")
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    row_iter = worksheet.iter_rows(values_only=True)

    try:
        header_row = next(row_iter)
    except StopIteration:
        workbook.close()
        return pd.DataFrame()

    headers: List[str] = []
    for index, cell in enumerate(header_row):
        name = str(cell).strip() if cell is not None else ""
        headers.append(name if name else f"列{index + 1}")

    estimated_total = max((worksheet.max_row or 1) - 1, 1)
    # 使用列表按列索引存储，避免重复表头（如两个「序号」）导致列长度不一致
    column_data: List[List] = [[] for _ in headers]
    row_count = 0

    for row in row_iter:
        row_count += 1
        for index in range(len(headers)):
            value = row[index] if index < len(row) else None
            column_data[index].append(value)

        if row_count == 1 or row_count % 500 == 0:
            total_hint = max(estimated_total, row_count)
            percent = 0.05 + min(row_count / total_hint, 1.0) * 0.5
            callback(percent, f"正在读取数据 {row_count}/{total_hint} 行...")

    workbook.close()
    callback(0.55, f"读取完成，共 {row_count} 行")
    callback(0.56, "正在整理表格数据...")
    df = pd.DataFrame(column_data).T
    df.columns = pd.Index(headers, dtype=object)
    return df


def detect_template(df: pd.DataFrame) -> TemplateConfig:
    """根据表头自动识别 Excel 模板类型。"""
    columns = set(df.columns)

    internal_keys = {"销售渠道", "货品名称"}
    if internal_keys.issubset(columns):
        return INTERNAL_TEMPLATE

    external_keys = {"客户", "产品名称", "销售"}
    if external_keys.issubset(columns):
        return EXTERNAL_TEMPLATE

    format_hints = []
    for template in TEMPLATES:
        format_hints.append(
            f"{template.label}：{', '.join(template.required_columns)}"
        )

    raise AggregateError(
        "无法识别 Excel 格式，请确认表头包含以下列之一：\n"
        + "\n".join(format_hints)
    )


def _validate_columns(df: pd.DataFrame, template: TemplateConfig) -> None:
    """校验模板必需列是否存在。"""
    missing = [col for col in template.required_columns if col not in df.columns]
    if missing:
        raise AggregateError(
            f"「{template.label}」缺少必需列：{', '.join(missing)}"
        )


def _convert_numeric_column(df: pd.DataFrame, column: str) -> None:
    """将指定列转为数值，并收集无效行信息。"""
    converted = pd.to_numeric(df[column], errors="coerce")
    invalid_mask = converted.isna() & df[column].notna()
    invalid_mask = invalid_mask & (df[column].astype(str).str.strip() != "")

    if invalid_mask.any():
        # Excel 行号 = DataFrame 索引 + 2（含表头）
        invalid_rows = [int(idx) + 2 for idx in df.index[invalid_mask].tolist()[:10]]
        suffix = "..." if invalid_mask.sum() > 10 else ""
        raise AggregateError(
            f"列「{column}」存在无法转换为数字的数据，"
            f"示例行号：{', '.join(map(str, invalid_rows))}{suffix}"
        )

    df[column] = converted.fillna(0)


def _format_output_values(df: pd.DataFrame, sum_columns: List[str]) -> pd.DataFrame:
    """格式化输出数值：数量为整数时去小数，金额保留两位小数。"""
    result = df.copy()

    if "数量" in sum_columns and "数量" in result.columns:
        qty = result["数量"]
        result["数量"] = qty.apply(
            lambda value: int(value)
            if float(value).is_integer()
            else round(float(value), 4)
        )

    if "金额" in sum_columns and "金额" in result.columns:
        result["金额"] = result["金额"].apply(lambda value: round(float(value), 2))

    return result


def aggregate_excel(
    input_path: Union[str, Path],
    progress_callback: Optional[ProgressCallback] = None,
) -> AggregateResult:
    """
    读取 Excel 并自动识别模板后汇总数量与金额。

    :param input_path: 输入 Excel 路径
    :param progress_callback: 可选进度回调 (0~1, 状态说明)
    :return: 汇总结果
    """
    report = progress_callback or _noop_progress
    path = Path(input_path)
    if not path.exists():
        raise AggregateError(f"文件不存在：{path}")

    if path.suffix.lower() not in {".xlsx", ".xls"}:
        raise AggregateError("仅支持 .xlsx 或 .xls 格式的 Excel 文件")

    try:
        report(0.02, "准备读取 Excel 文件...")
        df = _read_excel_with_progress(path, report)
    except AggregateError:
        raise
    except Exception as exc:
        raise AggregateError(f"读取 Excel 失败：{exc}") from exc

    if df.empty:
        raise AggregateError("Excel 中没有数据行")

    report(0.57, "正在标准化表头...")
    df = _normalize_columns(df)

    report(0.58, "正在识别表格类型...")
    template = detect_template(df)
    _validate_columns(df, template)

    report(0.65, f"已识别：{template.label}，正在提取数据...")
    source_columns = list(
        dict.fromkeys(template.group_by + template.sum_columns + template.keep_columns)
    )
    work_df = df[source_columns].copy()
    raw_row_count = len(work_df)

    # 过滤名称列为空的行（含 NaN、None 字符串等）
    name_column = template.name_filter_column
    for text_column in template.group_by:
        work_df[text_column] = work_df[text_column].astype(str).str.strip()
    empty_markers = {"", "nan", "none", "null", "NaN", "None"}
    work_df = work_df[~work_df[name_column].isin(empty_markers)]

    if work_df.empty:
        raise AggregateError(f"过滤空「{name_column}」后没有可汇总的数据")

    report(0.72, "正在校验数量与金额...")
    for column in template.sum_columns:
        _convert_numeric_column(work_df, column)

    report(0.82, "正在按规则合并汇总...")
    agg_map = {column: "sum" for column in template.sum_columns}
    for column in template.keep_columns:
        agg_map[column] = "first"

    grouped = (
        work_df.groupby(template.group_by, as_index=False)
        .agg(agg_map)
        .loc[:, template.output_columns]
    )

    report(0.92, "正在整理汇总结果...")
    grouped = grouped.sort_values(
        by=template.group_by,
        kind="stable",
    ).reset_index(drop=True)

    grouped = _format_output_values(grouped, template.sum_columns)
    report(1.0, "汇总完成")

    return AggregateResult(
        data=grouped,
        raw_row_count=raw_row_count,
        result_row_count=len(grouped),
        template_id=template.id,
        template_label=template.label,
        dimension_label=template.dimension_label,
    )


def export_excel(result: AggregateResult, output_path: Union[str, Path]) -> None:
    """导出汇总结果到 Excel。"""
    path = Path(output_path)
    if path.suffix.lower() != ".xlsx":
        path = path.with_suffix(".xlsx")

    template = next(
        (item for item in TEMPLATES if item.id == result.template_id),
        INTERNAL_TEMPLATE,
    )
    output_columns = list(result.data.columns)
    excel_column_widths = template.excel_column_widths

    try:
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            result.data.to_excel(writer, index=False, sheet_name="汇总结果")
            worksheet = writer.sheets["汇总结果"]

            # 表头样式：灰色底、深绿色加粗居中文字
            header_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
            header_font = Font(name="Calibri", size=10, bold=True, color="006400")
            header_alignment = Alignment(horizontal="center", vertical="center")

            for index, column in enumerate(output_columns, start=1):
                cell = worksheet.cell(row=1, column=index)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

                column_letter = cell.column_letter
                worksheet.column_dimensions[column_letter].width = excel_column_widths.get(
                    column, 16
                )

            worksheet.row_dimensions[1].height = 18
            worksheet.auto_filter.ref = worksheet.dimensions
            worksheet.freeze_panes = "A2"
    except Exception as exc:
        raise AggregateError(f"导出 Excel 失败：{exc}") from exc
